import openpyxl
import json

wb = openpyxl.load_workbook('C:/Users/l5857/TCC - LETICIA/MAPEAMENTO DE CAMPO PRÁTICO (1).xlsx')

# Read MARÇO sheet fully
ws = wb['MARÇO']
print(f'=== Sheet: MARÇO ===')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
for i, row in enumerate(ws.iter_rows(values_only=True)):
    non_none = [x for x in row if x is not None]
    if non_none:
        print(f'Row {i+1}: {row}')
