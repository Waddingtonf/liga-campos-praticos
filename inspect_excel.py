import openpyxl

wb = openpyxl.load_workbook('C:/Users/l5857/TCC - LETICIA/MAPEAMENTO DE CAMPO PRÁTICO (1).xlsx', data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if len(row) > 7:
            val = row[7]
            if val and isinstance(val, str) and '{' in val:
                print(f"Sheet: {sheet_name}, Row {r_idx+1}, Col H: {val}")
