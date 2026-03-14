import openpyxl
import json
import re
from collections import defaultdict

wb = openpyxl.load_workbook('C:/Users/l5857/TCC - LETICIA/MAPEAMENTO DE CAMPO PRÁTICO (1).xlsx')

def parse_alunos(val):
    if not val or not isinstance(val, str):
        return []
    val = val.strip()
    if val.startswith('{') and val.endswith('}'):
        # Ex: {"João"; "Maria"}
        inner = val[1:-1].strip()
        # Split by ;
        parts = inner.split(';')
        # Clean up quotes and whitespace
        return [p.replace('"', '').strip() for p in parts if p.strip()]
    return []

all_sheets_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    # Header is at row 2 (index 1)
    data = []
    for i, row in enumerate(rows):
        if i < 2:
            continue
        if row[0] is None and row[1] is None:
            continue
        
        raw_h = str(row[7]).strip() if row[7] else ''
        alunos_list = parse_alunos(raw_h)
        
        entry = {
            'unidade': str(row[0]).strip() if row[0] else '',
            'setor': str(row[1]).strip() if row[1] else '',
            'turno': str(row[2]).strip() if row[2] else '',
            'profissional': str(row[3]).strip() if row[3] else '',
            'capacidade': float(row[4]) if row[4] is not None else 0,
            'ocupacao': float(row[5]) if row[5] is not None else None,
            'curso': str(row[6]).strip() if row[6] else '',
            'tipo_ocupacao': raw_h if not alunos_list else 'ALUNOS LISTADOS',
            'alunos': alunos_list,
            'periodo': str(row[8]).strip() if row[8] else '',
        }
        if entry['unidade'] in ('UNIDADE', ''):
            continue
        data.append(entry)
    
    all_sheets_data[sheet_name] = data
    print(f"Sheet '{sheet_name}': {len(data)} records")

# Write JSON
with open('C:/Users/l5857/TCC - LETICIA/data.json', 'w', encoding='utf-8') as f:
    json.dump(all_sheets_data, f, ensure_ascii=False, indent=2)

print("\ndata.json written successfully!")
print("\n=== MARÇO Summary ===")
marco = all_sheets_data.get('MARÇO', [])
unidades = defaultdict(set)
for r in marco:
    unidades[r['unidade']].add(r['setor'])
for u, setores in sorted(unidades.items()):
    print(f"\n{u} ({len(setores)} setores):")
    for s in sorted(setores):
        print(f"  - {s}")
