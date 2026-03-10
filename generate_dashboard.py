import openpyxl
import json

# ─────────────────────────────────────────────────────────────────
# 1. READ EXCEL
# ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook('C:/Users/l5857/TCC - LETICIA/MAPEAMENTO DE CAMPO PRÁTICO (1).xlsx')

all_data = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # skip title + header
            continue
        if not row[0] or row[0] == 'UNIDADE':
            continue
        records.append({
            'unidade':      str(row[0]).strip() if row[0] else '',
            'setor':        str(row[1]).strip() if row[1] else '',
            'turno':        str(row[2]).strip() if row[2] else '',
            'profissional': str(row[3]).strip() if row[3] else '',
            'capacidade':   float(row[4]) if row[4] is not None else 0,
            'ocupacao':     float(row[5]) if row[5] is not None else None,
            'curso':        str(row[6]).strip() if row[6] else '',
            'tipo_ocupacao':str(row[7]).strip() if row[7] else '',
            'periodo':      str(row[8]).strip() if row[8] else '',
        })
    all_data[sheet_name] = records

data_json = json.dumps(all_data, ensure_ascii=False)

# ──────────────────────────────────────────────────────────────
# 2. Generate HTML
# ──────────────────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Mapeamento de Campos Práticos</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --cecan: #8b5cf6;
    --hla:   #3b82f6;
    --pol:   #f59e0b;
    --full:  #ef4444;
    --partial:#f97316;
    --avail: #22c55e;
    --nodata:#9ca3af;
    --bg:    #0f172a;
    --card:  #1e293b;
    --border:#334155;
    --text:  #f1f5f9;
    --muted: #94a3b8;
  }}
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:var(--bg); color:var(--text); font-family:'Segoe UI',system-ui,sans-serif; min-height:100vh; }}
  
  /* ── HEADER ── */
  .header {{
    background: linear-gradient(135deg, #1e1b4b 0%, #312e81 50%, #1e3a5f 100%);
    padding: 20px 32px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 1px solid rgba(255,255,255,.1);
    box-shadow: 0 4px 24px rgba(0,0,0,.4);
  }}
  .header-left {{ display:flex; align-items:center; gap:16px; }}
  .logo-circle {{
    width:52px; height:52px; border-radius:50%;
    background: linear-gradient(135deg, #6366f1, #8b5cf6);
    display:flex; align-items:center; justify-content:center;
    font-size:24px; box-shadow:0 0 20px rgba(139,92,246,.5);
  }}
  .header-title h1 {{ font-size:1.4rem; font-weight:700; letter-spacing:.3px; }}
  .header-title p {{ font-size:.8rem; color:#a5b4fc; margin-top:2px; }}
  .header-right {{ display:flex; align-items:center; gap:12px; }}
  .date-badge {{
    background:rgba(255,255,255,.08);
    border:1px solid rgba(255,255,255,.12);
    border-radius:8px; padding:6px 14px;
    font-size:.8rem; color:#c7d2fe;
  }}
  .google-badge {{
    background: linear-gradient(135deg,rgba(66,133,244,.2),rgba(52,168,83,.2));
    border:1px solid rgba(66,133,244,.3);
    border-radius:8px; padding:6px 14px;
    font-size:.75rem; color:#93c5fd;
    display:flex; align-items:center; gap:6px;
  }}
  
  /* ── MAIN CONTENT ── */
  .main {{ padding:24px 32px; }}
  
  /* ── SEARCH BAR ── */
  .search-wrap {{ position:relative; margin-bottom:20px; }}
  .search-input {{
    width:100%; padding:10px 16px 10px 40px;
    background:var(--card); border:1px solid var(--border);
    border-radius:10px; color:var(--text); font-size:.9rem;
    outline:none; transition:border-color .2s;
  }}
  .search-input::placeholder {{ color:var(--muted); }}
  .search-input:focus {{ border-color:#6366f1; }}
  .search-icon {{ position:absolute; left:12px; top:50%; transform:translateY(-50%); opacity:.5; font-size:1rem; }}
  .search-clear {{ position:absolute; right:12px; top:50%; transform:translateY(-50%); cursor:pointer; opacity:.5; font-size:.8rem; padding:4px; background:none; border:none; color:var(--text); }}
  .no-results {{ text-align:center; padding:40px; color:var(--muted); }}

  /* ── MONTH TABS ── */
  .month-tabs {{ display:flex; gap:8px; margin-bottom:24px; flex-wrap:wrap; }}
  .month-tab {{
    padding:8px 20px; border-radius:8px; border:1px solid var(--border);
    background:var(--card); color:var(--muted); cursor:pointer;
    font-size:.85rem; font-weight:500; transition:all .2s;
  }}
  .month-tab:hover {{ border-color:#6366f1; color:var(--text); }}
  .month-tab.active {{
    background: linear-gradient(135deg,#4f46e5,#7c3aed);
    border-color:#6366f1; color:#fff;
    box-shadow:0 0 16px rgba(99,102,241,.4);
  }}
  
  /* ── KPI CARDS ── */
  .kpi-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(170px,1fr)); gap:14px; margin-bottom:28px; }}
  .kpi-card {{
    background:var(--card); border:1px solid var(--border);
    border-radius:14px; padding:18px 20px;
    display:flex; flex-direction:column; gap:6px;
    transition:transform .2s, box-shadow .2s;
  }}
  .kpi-card:hover {{ transform:translateY(-2px); box-shadow:0 8px 24px rgba(0,0,0,.3); }}
  .kpi-card .kpi-icon {{ font-size:1.6rem; }}
  .kpi-card .kpi-value {{ font-size:2rem; font-weight:700; line-height:1; }}
  .kpi-card .kpi-label {{ font-size:.75rem; color:var(--muted); }}
  .kpi-card.red .kpi-value {{ color:var(--full); }}
  .kpi-card.orange .kpi-value {{ color:var(--partial); }}
  .kpi-card.green .kpi-value {{ color:var(--avail); }}
  .kpi-card.purple .kpi-value {{ color:#a78bfa; }}
  .kpi-card.blue .kpi-value {{ color:#60a5fa; }}
  .kpi-card.amber .kpi-value {{ color:#fbbf24; }}
  
  /* ── UNIT TABS ── */
  .unit-tabs {{ display:flex; gap:8px; margin-bottom:20px; flex-wrap:wrap; align-items:center; }}
  .unit-tab {{
    padding:8px 22px; border-radius:24px; border:2px solid transparent;
    cursor:pointer; font-size:.85rem; font-weight:600; transition:all .2s;
    background:var(--card);
  }}
  .unit-tab[data-unit="ALL"] {{ border-color:var(--border); color:var(--text); }}
  .unit-tab[data-unit="CECAN"] {{ color:var(--cecan); border-color:var(--cecan); }}
  .unit-tab[data-unit="HLA"]   {{ color:var(--hla);   border-color:var(--hla);   }}
  .unit-tab[data-unit="POL"]   {{ color:var(--pol);   border-color:var(--pol);   }}
  .unit-tab.active[data-unit="ALL"]   {{ background:#475569; color:#fff; }}
  .unit-tab.active[data-unit="CECAN"] {{ background:var(--cecan); color:#fff; box-shadow:0 0 16px rgba(139,92,246,.4); }}
  .unit-tab.active[data-unit="HLA"]   {{ background:var(--hla);   color:#fff; box-shadow:0 0 16px rgba(59,130,246,.4); }}
  .unit-tab.active[data-unit="POL"]   {{ background:var(--pol);   color:#fff; box-shadow:0 0 16px rgba(245,158,11,.4); }}
  
  /* ── LEGEND ── */
  .legend {{ display:flex; align-items:center; gap:16px; margin-left:auto; flex-wrap:wrap; }}
  .legend-item {{ display:flex; align-items:center; gap:6px; font-size:.75rem; color:var(--muted); }}
  .legend-dot {{ width:10px; height:10px; border-radius:50%; }}
  
  /* ── UNIT SECTION ── */
  .unit-section {{ margin-bottom:36px; }}
  .unit-header {{
    display:flex; align-items:center; gap:12px;
    margin-bottom:16px; padding-bottom:10px;
    border-bottom:1px solid var(--border);
  }}
  .unit-badge {{
    padding:4px 14px; border-radius:20px;
    font-size:1rem; font-weight:700; letter-spacing:.5px;
  }}
  .unit-badge.CECAN {{ background:rgba(139,92,246,.2); color:var(--cecan); border:1px solid rgba(139,92,246,.3); }}
  .unit-badge.HLA   {{ background:rgba(59,130,246,.2);  color:var(--hla);   border:1px solid rgba(59,130,246,.3); }}
  .unit-badge.POL   {{ background:rgba(245,158,11,.2);  color:var(--pol);   border:1px solid rgba(245,158,11,.3); }}
  .unit-subtitle {{ font-size:.85rem; color:var(--muted); }}
  
  /* ── SECTOR GRID ── */
  .sector-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(200px,1fr)); gap:12px; }}
  
  /* ── SECTOR CARD ── */
  .sector-card {{
    background:var(--card); border:1px solid var(--border);
    border-radius:14px; padding:16px;
    cursor:pointer; position:relative; overflow:hidden;
    transition:transform .2s, box-shadow .2s, border-color .2s;
  }}
  .sector-card::before {{
    content:''; position:absolute; top:0; left:0; right:0; height:4px;
    border-radius:14px 14px 0 0;
  }}
  .sector-card.status-full::before    {{ background:var(--full); }}
  .sector-card.status-partial::before {{ background:var(--partial); }}
  .sector-card.status-avail::before   {{ background:var(--avail); }}
  .sector-card.status-nodata::before  {{ background:var(--nodata); }}
  .sector-card:hover {{
    transform:translateY(-4px);
    box-shadow:0 12px 32px rgba(0,0,0,.4);
  }}
  .sector-card.status-full:hover    {{ border-color:var(--full);    box-shadow:0 12px 32px rgba(239,68,68,.2);  }}
  .sector-card.status-partial:hover {{ border-color:var(--partial); box-shadow:0 12px 32px rgba(249,115,22,.2); }}
  .sector-card.status-avail:hover   {{ border-color:var(--avail);   box-shadow:0 12px 32px rgba(34,197,94,.2);  }}
  
  .sector-name {{
    font-size:.88rem; font-weight:700; margin-bottom:10px;
    line-height:1.3; min-height:2.2em;
  }}
  
  /* occupation bar */
  .occ-bar-wrap {{ height:6px; background:rgba(255,255,255,.08); border-radius:3px; overflow:hidden; margin-bottom:8px; }}
  .occ-bar {{ height:100%; border-radius:3px; transition:width .4s ease; }}
  .status-full    .occ-bar {{ background:var(--full); }}
  .status-partial .occ-bar {{ background:var(--partial); }}
  .status-avail   .occ-bar {{ background:var(--avail); }}
  .status-nodata  .occ-bar {{ background:var(--nodata); }}
  
  .occ-numbers {{ display:flex; align-items:center; justify-content:space-between; }}
  .occ-text {{ font-size:.78rem; font-weight:700; }}
  .cap-text {{ font-size:.72rem; color:var(--muted); }}
  .status-badge {{
    font-size:.68rem; font-weight:700; padding:2px 8px;
    border-radius:12px; text-transform:uppercase; letter-spacing:.5px;
  }}
  .status-full    .status-badge {{ background:rgba(239,68,68,.2);    color:var(--full);    border:1px solid rgba(239,68,68,.3); }}
  .status-partial .status-badge {{ background:rgba(249,115,22,.2);   color:var(--partial); border:1px solid rgba(249,115,22,.3); }}
  .status-avail   .status-badge {{ background:rgba(34,197,94,.2);    color:var(--avail);   border:1px solid rgba(34,197,94,.3); }}
  .status-nodata  .status-badge {{ background:rgba(156,163,175,.2);  color:var(--nodata);  border:1px solid rgba(156,163,175,.3); }}
  
  /* ── ACTIVE PERIOD BADGE ── */
  .active-period-badge {{
    display:inline-flex; align-items:center; gap:4px;
    font-size:.65rem; padding:2px 7px; border-radius:10px;
    background:rgba(34,197,94,.15); color:#4ade80;
    border:1px solid rgba(34,197,94,.3);
    margin-top:6px; width:fit-content;
    animation: pulseGreen 2s infinite;
  }}
  .active-period-badge .dot {{ width:5px; height:5px; border-radius:50%; background:#4ade80; }}
  @keyframes pulseGreen {{
    0%,100% {{ box-shadow:0 0 0 0 rgba(34,197,94,.3); }}
    50% {{ box-shadow:0 0 0 4px rgba(34,197,94,0); }}
  }}
  .remaining-slots {{
    font-size:.7rem; color:var(--avail); font-weight:600; margin-top:4px;
  }}

  .prof-tags {{ margin-top:8px; display:flex; flex-wrap:wrap; gap:4px; }}
  .prof-tag {{
    font-size:.65rem; padding:2px 7px; border-radius:10px;
    background:rgba(255,255,255,.06); color:var(--muted);
    border:1px solid rgba(255,255,255,.08);
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis; max-width:100%;
  }}
  
  /* ── TOOLTIP ── */
  #tooltip {{
    position:fixed; z-index:9999; pointer-events:none;
    background:#0f172a; border:1px solid #334155;
    border-radius:14px; padding:16px 18px;
    max-width:340px; min-width:240px;
    box-shadow:0 20px 60px rgba(0,0,0,.6);
    opacity:0; transition:opacity .15s;
    font-size:.82rem; line-height:1.5;
  }}
  #tooltip.visible {{ opacity:1; }}
  .tt-header {{ display:flex; align-items:center; justify-content:space-between; margin-bottom:10px; gap:8px; }}
  .tt-setor {{ font-size:1rem; font-weight:700; }}
  .tt-unit {{ font-size:.72rem; padding:2px 8px; border-radius:10px; }}
  .tt-divider {{ border:none; border-top:1px solid var(--border); margin:8px 0; }}
  .tt-row {{ display:flex; align-items:flex-start; gap:6px; margin:4px 0; }}
  .tt-icon {{ width:16px; flex-shrink:0; opacity:.6; }}
  .tt-key {{ color:var(--muted); min-width:80px; }}
  .tt-val {{ font-weight:600; flex:1; }}
  .tt-turno-block {{ margin:6px 0; padding:8px; border-radius:8px; background:rgba(255,255,255,.04); border:1px solid var(--border); }}
  .tt-turno-header {{ display:flex; justify-content:space-between; align-items:center; font-weight:700; margin-bottom:4px; font-size:.78rem; }}
  .tt-turno-name {{ color:#a5b4fc; }}
  .tt-bar {{ height:5px; border-radius:3px; margin:4px 0; }}
  
  /* ── CHARTS SECTION ── */
  .charts-section {{
    display:grid; grid-template-columns:repeat(auto-fit,minmax(320px,1fr));
    gap:20px; margin-top:32px;
  }}
  .chart-card {{
    background:var(--card); border:1px solid var(--border);
    border-radius:16px; padding:22px;
  }}
  .chart-title {{ font-size:.95rem; font-weight:700; margin-bottom:16px; display:flex; align-items:center; gap:8px; }}
  .chart-title .icon {{ opacity:.7; }}
  .chart-wrap {{ position:relative; }}
  
  /* ── FOOTER ── */
  .footer {{ text-align:center; padding:24px; color:var(--muted); font-size:.75rem; border-top:1px solid var(--border); margin-top:32px; }}
  
  /* ── RESPONSIVE ── */
  @media(max-width:640px) {{
    .main {{ padding:16px; }}
    .header {{ padding:14px 16px; }}
    .header-title h1 {{ font-size:1.1rem; }}
    .kpi-grid {{ grid-template-columns:repeat(2,1fr); }}
    .legend {{ display:none; }}
  }}
</style>
</head>
<body>

<!-- HEADER -->
<header class="header">
  <div class="header-left">
    <div class="logo-circle">🏥</div>
    <div class="header-title">
      <h1>Mapeamento de Campos Práticos</h1>
      <p>Gestão de Setores e Ocupação por Unidade de Atuação</p>
    </div>
  </div>
  <div class="header-right">
    <div class="date-badge">📅 Março 2026</div>
    <div class="google-badge">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none"><path d="M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm-1 14H9V8h2v8zm4 0h-2V8h2v8z" fill="#4285F4"/></svg>
      Google Sheets Ready
    </div>
  </div>
</header>

<!-- MAIN -->
<main class="main">

  <!-- SEARCH BAR -->
  <div class="search-wrap">
    <span class="search-icon">🔍</span>
    <input class="search-input" id="searchInput" type="text" placeholder="Buscar setor por nome..."/>
    <button class="search-clear" id="searchClear" onclick="clearSearch()">✕</button>
  </div>

  <!-- MONTH TABS -->
  <div class="month-tabs" id="monthTabs">
    <button class="month-tab" data-sheet="JANEIRO(PLANEJAMENTO)">📋 Janeiro (Planejamento)</button>
    <button class="month-tab" data-sheet="FEVEREIRO (PLANEJAMENTO)">📋 Fevereiro (Planejamento)</button>
    <button class="month-tab active" data-sheet="MARÇO">✅ Março (Consolidado)</button>
  </div>

  <!-- KPI GRID -->
  <div class="kpi-grid" id="kpiGrid"></div>

  <!-- UNIT + LEGEND TOOLBAR -->
  <div class="unit-tabs" id="unitTabs">
    <button class="unit-tab active" data-unit="ALL">🏢 Todas as Unidades</button>
    <button class="unit-tab" data-unit="CECAN">CECAN</button>
    <button class="unit-tab" data-unit="HLA">HLA</button>
    <button class="unit-tab" data-unit="POL">POL</button>
    <div class="legend">
      <div class="legend-item"><div class="legend-dot" style="background:var(--full)"></div>Lotado</div>
      <div class="legend-item"><div class="legend-dot" style="background:var(--partial)"></div>Parcial</div>
      <div class="legend-item"><div class="legend-dot" style="background:var(--avail)"></div>Disponível</div>
      <div class="legend-item"><div class="legend-dot" style="background:var(--nodata)"></div>Sem dados</div>
    </div>
  </div>

  <!-- SECTOR MAP -->
  <div id="sectorMap"></div>

  <!-- CHARTS -->
  <div class="charts-section">
    <div class="chart-card">
      <div class="chart-title"><span class="icon">🍩</span> Distribuição de Ocupação</div>
      <div class="chart-wrap" style="max-height:260px"><canvas id="chartStatus"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title"><span class="icon">📊</span> Ocupação por Unidade</div>
      <div class="chart-wrap" style="max-height:260px"><canvas id="chartUnits"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title"><span class="icon">👥</span> Profissionais por Unidade</div>
      <div class="chart-wrap" style="max-height:260px"><canvas id="chartProfs"></canvas></div>
    </div>
  </div>

</main>

<!-- FOOTER -->
<footer class="footer">
  Sistema de Mapeamento de Campos Práticos &bull; Dados extraídos do MAPEAMENTO DE CAMPO PRÁTICO.xlsx &bull; Compatível com Google Sheets
</footer>

<!-- TOOLTIP -->
<div id="tooltip"></div>

<!-- ═══════════════════════════════════════════════════════ -->
<!-- DATA + SCRIPTS                                         -->
<!-- ═══════════════════════════════════════════════════════ -->
<script>
const RAW_DATA = {data_json};

// ────────────────────────────────────────────────────────────
// STATE
// ────────────────────────────────────────────────────────────
let currentSheet = 'MARÇO';
let currentUnit  = 'ALL';
let searchTerm   = '';
let charts = {{}};
const TODAY = new Date(2026, 2, 9); // March 9, 2026

// ────────────────────────────────────────────────────────────
// PERIOD PARSING
// ────────────────────────────────────────────────────────────
function parseBrDate(s) {{
  // Handles formats like "03/03/2026" or "3/3/2026"
  if (!s) return null;
  s = s.trim();
  const m = s.match(/(\\d{{1,2}})\\/(\\d{{1,2}})\\/(\\d{{4}})/);
  if (!m) return null;
  return new Date(parseInt(m[3]), parseInt(m[2])-1, parseInt(m[1]));
}}

function isPeriodActive(periodo) {{
  if (!periodo || periodo === 'PENDENTE') return false;
  // Try to split on ' - ' or '-'
  const sep = periodo.includes(' - ') ? ' - ' : '-';
  const parts = periodo.split(sep).map(p=>p.trim());
  if (parts.length < 2) return false;
  const start = parseBrDate(parts[0]);
  const end   = parseBrDate(parts[parts.length-1]);
  if (!start || !end) return false;
  return TODAY >= start && TODAY <= end;
}}

function sectorHasActivePeriod(rows) {{
  return rows.some(r => r.periodo && isPeriodActive(r.periodo));
}}

// ────────────────────────────────────────────────────────────
// UTILS
// ────────────────────────────────────────────────────────────
function getSheetData() {{
  return RAW_DATA[currentSheet] || [];
}}

function getUnitColor(u) {{
  return {{CECAN:'#8b5cf6',HLA:'#3b82f6',POL:'#f59e0b'}}[u] || '#6b7280';
}}

function abbreviateProfissional(p) {{
  const map = {{
    'TÉCNICO DE ENFERMAGEM':'TEC.ENF',
    'TÉCNICO DE RADIOLOGIA':'TEC.RAD',
    'ENFERMEIRO':'ENF',
    'ENFERMAGEM':'ENF',
    'FARMACÊUTICO':'FARM',
    'FARMÁCIA':'FARM',
    'BIOMEDICINA':'BIOMED',
    'NUTRICIONISTA':'NUTRI',
    'NUTRIÇÃO':'NUTRI',
    'FISIOTERAPEUTA':'FISIO',
    'ASSISTENTE SOCIAL':'ASS.SOC',
    'PSICOLOGIA':'PSICO',
    'DENTISTA':'DENT',
  }};
  if (map[p]) return map[p];
  return p.length > 10 ? p.substring(0,9)+'…' : p;
}}

// ────────────────────────────────────────────────────────────
// SECTOR GROUPING
// ────────────────────────────────────────────────────────────
function groupBySector(data) {{
  // Map: unidade|setor → {{ setor, unidade, profissionais: [[turno, prof, cap, occ, curso, tipo, periodo],...] }}
  const map = new Map();
  data.forEach(r => {{
    if (!r.unidade || !r.setor) return;
    const key = r.unidade + '|||' + r.setor;
    if (!map.has(key)) map.set(key, {{ setor:r.setor, unidade:r.unidade, rows:[] }});
    map.get(key).rows.push(r);
  }});
  return Array.from(map.values());
}}

function getSectorStatus(rows) {{
  const validRows = rows.filter(r => r.capacidade > 0);
  if (!validRows.length) return 'nodata';
  const totalCap = validRows.reduce((s,r) => s + r.capacidade, 0);
  const knownOcc = validRows.filter(r => r.ocupacao !== null && r.ocupacao !== undefined);
  if (!knownOcc.length) return 'nodata';
  const totalOcc = knownOcc.reduce((s,r) => s + (r.ocupacao || 0), 0);
  if (totalOcc === 0) return 'avail';
  if (totalOcc < totalCap) return 'partial';
  return 'full';
}}

function getSectorTotals(rows) {{
  const validRows = rows.filter(r => r.capacidade > 0);
  const totalCap = validRows.reduce((s,r) => s + r.capacidade, 0);
  const totalOcc = validRows.reduce((s,r) => s + (r.ocupacao || 0), 0);
  return {{ totalCap, totalOcc, pct: totalCap > 0 ? Math.min(100, Math.round(totalOcc/totalCap*100)) : 0 }};
}}

function getUniqueProfissionais(rows) {{
  return [...new Set(rows.map(r=>r.profissional).filter(Boolean))];
}}

// ────────────────────────────────────────────────────────────
// KPI
// ────────────────────────────────────────────────────────────
function renderKPI() {{
  const data = getSheetData();
  const units = currentUnit === 'ALL' ? data : data.filter(r=>r.unidade===currentUnit);
  const sectors = groupBySector(units);
  
  const counts = {{full:0,partial:0,avail:0,nodata:0}};
  sectors.forEach(s => counts[getSectorStatus(s.rows)]++);
  const totalCap = units.reduce((s,r)=>s+(r.capacidade||0),0);
  const totalOcc = units.reduce((s,r)=>s+(r.ocupacao||0),0);
  
  document.getElementById('kpiGrid').innerHTML = `
    <div class="kpi-card purple"><div class="kpi-icon">🗂️</div><div class="kpi-value">${{sectors.length}}</div><div class="kpi-label">Total de Setores</div></div>
    <div class="kpi-card red"><div class="kpi-icon">🔴</div><div class="kpi-value">${{counts.full}}</div><div class="kpi-label">Setores Lotados</div></div>
    <div class="kpi-card orange"><div class="kpi-icon">🟠</div><div class="kpi-value">${{counts.partial}}</div><div class="kpi-label">Parcialmente Ocupados</div></div>
    <div class="kpi-card green"><div class="kpi-icon">🟢</div><div class="kpi-value">${{counts.avail}}</div><div class="kpi-label">Disponíveis</div></div>
    <div class="kpi-card blue"><div class="kpi-icon">💺</div><div class="kpi-value">${{totalCap.toFixed(0)}}</div><div class="kpi-label">Cap. Total (vagas)</div></div>
    <div class="kpi-card amber"><div class="kpi-icon">📌</div><div class="kpi-value">${{totalOcc.toFixed(0)}}</div><div class="kpi-label">Vagas Ocupadas</div></div>
  `;
}}

// ────────────────────────────────────────────────────────────
// SECTOR MAP
// ────────────────────────────────────────────────────────────
function renderMap() {{
  const data = getSheetData();
  const unitOrder = ['CECAN','HLA','POL'];
  const activeUnits = currentUnit === 'ALL' ? unitOrder : [currentUnit];
  const term = searchTerm.toLowerCase();
  
  let html = '';
  activeUnits.forEach(unit => {{
    const unitData = data.filter(r=>r.unidade===unit);
    if (!unitData.length) return;
    let sectors = groupBySector(unitData);
    if (term) sectors = sectors.filter(s => s.setor.toLowerCase().includes(term));
    const unitColor = getUnitColor(unit);
    
    html += `<div class="unit-section" data-unit="${{unit}}">
      <div class="unit-header">
        <span class="unit-badge ${{unit}}">${{unit}}</span>
        <span class="unit-subtitle">${{sectors.length}} setores</span>
      </div>
      <div class="sector-grid">`;
    
    sectors.forEach(s => {{
      const status = getSectorStatus(s.rows);
      const {{ totalCap, totalOcc, pct }} = getSectorTotals(s.rows);
      const profs = getUniqueProfissionais(s.rows);
      const statusLabel = {{full:'Lotado',partial:'Parcial',avail:'Disponível',nodata:'Sem dados'}}[status];
      const barWidth = status === 'nodata' ? 0 : pct;
      
      const profTags = profs.slice(0,3).map(p=>
        `<span class="prof-tag">${{abbreviateProfissional(p)}}</span>`
      ).join('');
      
      const tooltipData = encodeURIComponent(JSON.stringify({{
        setor: s.setor, unidade: s.unidade, rows: s.rows
      }}));
      
      const isActive = sectorHasActivePeriod(s.rows);
      const remaining = Math.max(0, totalCap - totalOcc);
      const activeBadge = isActive ? `<div class="active-period-badge"><div class="dot"></div>ATIVO EM MARÇO</div>` : '';
      const remainSlots = (status==='partial' || status==='avail') && remaining>0
        ? `<div class="remaining-slots">+${{remaining.toFixed(0)}} vaga(s) livre(s)</div>` : '';
      
      html += `<div class="sector-card status-${{status}}" 
                    data-tooltip="${{tooltipData}}"
                    onmouseenter="showTooltip(event,this)"
                    onmouseleave="hideTooltip()"
                    onmousemove="moveTooltip(event)">
        <div class="sector-name">${{s.setor}}</div>
        <div class="occ-bar-wrap"><div class="occ-bar" style="width:${{barWidth}}%"></div></div>
        <div class="occ-numbers">
          <div>
            <span class="occ-text">${{totalOcc.toFixed(0)}}</span>
            <span class="cap-text"> / ${{totalCap.toFixed(0)}} vagas</span>
          </div>
          <span class="status-badge">${{statusLabel}}</span>
        </div>
        ${{activeBadge}}
        ${{remainSlots}}
        <div class="prof-tags">${{profTags}}${{profs.length>3?`<span class="prof-tag">+${{profs.length-3}}</span>`:''}}</div>
      </div>`;
    }});
    
    html += `</div></div>`;
  }});
  
  const noResult = term && !html.includes('sector-card') ? `<div class="no-results">🔍 Nenhum setor encontrado para "${{searchTerm}}"</div>` : '';
  document.getElementById('sectorMap').innerHTML = (noResult || html) || '<p style="color:var(--muted);text-align:center;padding:40px">Nenhum dado encontrado para este mês.</p>';
}}

// ────────────────────────────────────────────────────────────
// TOOLTIP
// ────────────────────────────────────────────────────────────
const tooltip = document.getElementById('tooltip');
let tooltipX = 0, tooltipY = 0;

function showTooltip(e, el) {{
  const raw = JSON.parse(decodeURIComponent(el.dataset.tooltip));
  const uc = getUnitColor(raw.unidade);
  
  // Group rows by turno+profissional
  const grouped = new Map();
  raw.rows.forEach(r => {{
    const key = r.turno + '||' + r.profissional;
    if (!grouped.has(key)) grouped.set(key, r);
  }});
  
  // Build turno blocks
  const turnoMap = new Map();
  raw.rows.forEach(r => {{
    if (!turnoMap.has(r.turno)) turnoMap.set(r.turno, []);
    turnoMap.get(r.turno).push(r);
  }});
  
  let turnoBlocks = '';
  turnoMap.forEach((rows, turno) => {{
    const tCap = rows.reduce((s,r)=>s+(r.capacidade||0),0);
    const tOcc = rows.reduce((s,r)=>s+(r.ocupacao||0),0);
    const tPct = tCap > 0 ? Math.min(100,Math.round(tOcc/tCap*100)) : 0;
    const tStatus = rows[0] ? getSectorStatus(rows) : 'nodata';
    const tColor = {{full:'#ef4444',partial:'#f97316',avail:'#22c55e',nodata:'#9ca3af'}}[tStatus];
    
    const profRows = rows.map(r => {{
      const occ = r.ocupacao !== null && r.ocupacao !== undefined ? r.ocupacao : '—';
      const pct2 = r.capacidade>0 && r.ocupacao!==null ? Math.min(100,Math.round((r.ocupacao/r.capacidade)*100)) : 0;
      const extra = r.curso ? `<div style="font-size:.7rem;color:#a5b4fc;margin-top:2px">📚 ${{r.curso}}</div>` : '';
      const periodo = r.periodo ? `<div style="font-size:.7rem;color:#94a3b8;margin-top:1px">📅 ${{r.periodo}}</div>` : '';
      return `<div style="margin-bottom:6px">
        <div style="display:flex;justify-content:space-between;align-items:center">
          <span style="font-size:.72rem;color:#e2e8f0">${{r.profissional}}</span>
          <span style="font-size:.72rem;font-weight:700;color:${{tColor}}">${{occ}}/${{r.capacidade}}</span>
        </div>
        <div style="height:3px;background:rgba(255,255,255,.08);border-radius:2px;margin-top:3px">
          <div style="height:100%;width:${{pct2}}%;background:${{tColor}};border-radius:2px"></div>
        </div>
        ${{extra}}${{periodo}}
      </div>`;
    }}).join('');
    
    turnoBlocks += `<div class="tt-turno-block">
      <div class="tt-turno-header">
        <span class="tt-turno-name">⏰ ${{turno || 'N/A'}}</span>
        <span style="font-size:.7rem;color:${{tColor}}">${{tOcc.toFixed(0)}}/${{tCap.toFixed(0)}} (${{tPct}}%)</span>
      </div>
      ${{profRows}}
    </div>`;
  }});
  
  tooltip.innerHTML = `
    <div class="tt-header">
      <span class="tt-setor">${{raw.setor}}</span>
      <span class="tt-unit" style="background:${{uc}}22;color:${{uc}};border:1px solid ${{uc}}44">${{raw.unidade}}</span>
    </div>
    <hr class="tt-divider"/>
    ${{turnoBlocks}}
  `;
  
  moveTooltip(e);
  tooltip.classList.add('visible');
}}

function hideTooltip() {{
  tooltip.classList.remove('visible');
}}

function moveTooltip(e) {{
  const pad = 16;
  const tw = tooltip.offsetWidth || 300;
  const th = tooltip.offsetHeight || 200;
  let x = e.clientX + pad;
  let y = e.clientY + pad;
  if (x + tw > window.innerWidth - pad) x = e.clientX - tw - pad;
  if (y + th > window.innerHeight - pad) y = e.clientY - th - pad;
  tooltip.style.left = x + 'px';
  tooltip.style.top  = y + 'px';
}}

// ────────────────────────────────────────────────────────────
// CHARTS
// ────────────────────────────────────────────────────────────
function destroyCharts() {{
  Object.values(charts).forEach(c => c && c.destroy());
  charts = {{}};
}}

function renderCharts() {{
  destroyCharts();
  const data = getSheetData();
  const units = currentUnit === 'ALL' ? data : data.filter(r=>r.unidade===currentUnit);
  const sectors = groupBySector(units);
  
  // Chart 1: Status doughnut
  const counts = {{full:0,partial:0,avail:0,nodata:0}};
  sectors.forEach(s => counts[getSectorStatus(s.rows)]++);
  
  charts.status = new Chart(document.getElementById('chartStatus'), {{
    type: 'doughnut',
    data: {{
      labels: ['Lotado','Parcial','Disponível','Sem dados'],
      datasets: [{{
        data: [counts.full, counts.partial, counts.avail, counts.nodata],
        backgroundColor: ['#ef4444','#f97316','#22c55e','#9ca3af'],
        borderColor: '#1e293b', borderWidth: 3
      }}]
    }},
    options: {{
      responsive:true, maintainAspectRatio:true,
      cutout:'65%',
      plugins:{{
        legend:{{ position:'bottom', labels:{{ color:'#94a3b8', padding:12, font:{{size:11}} }} }},
        tooltip:{{ callbacks:{{ label: ctx => ` ${{ctx.label}}: ${{ctx.parsed}}` }} }}
      }}
    }}
  }});
  
  // Chart 2: Occupation % per unit
  const unitOrder = ['CECAN','HLA','POL'];
  const unitLabels = [], unitOccPct = [], unitColors = [];
  unitOrder.forEach(u => {{
    const ud = data.filter(r=>r.unidade===u);
    if (!ud.length) return;
    const cap = ud.reduce((s,r)=>s+(r.capacidade||0),0);
    const occ = ud.reduce((s,r)=>s+(r.ocupacao||0),0);
    unitLabels.push(u);
    unitOccPct.push(cap > 0 ? Math.round(occ/cap*100) : 0);
    unitColors.push(getUnitColor(u));
  }});
  
  charts.units = new Chart(document.getElementById('chartUnits'), {{
    type: 'bar',
    data: {{
      labels: unitLabels,
      datasets: [{{
        label: 'Ocupação (%)',
        data: unitOccPct,
        backgroundColor: unitColors.map(c=>c+'aa'),
        borderColor: unitColors, borderWidth:2, borderRadius:8
      }}]
    }},
    options: {{
      responsive:true, maintainAspectRatio:true,
      plugins:{{ legend:{{ display:false }}, tooltip:{{ callbacks:{{ label: ctx=>`${{ctx.parsed.y}}%` }} }} }},
      scales:{{
        y:{{ min:0, max:100, grid:{{ color:'#334155' }}, ticks:{{ color:'#94a3b8', callback:v=>v+'%' }} }},
        x:{{ grid:{{ display:false }}, ticks:{{ color:'#94a3b8' }} }}
      }}
    }}
  }});
  
  // Chart 3: Top profissionais
  const profCount = new Map();
  units.forEach(r => {{
    if (r.profissional) profCount.set(r.profissional, (profCount.get(r.profissional)||0)+1);
  }});
  const sorted = [...profCount.entries()].sort((a,b)=>b[1]-a[1]).slice(0,8);
  const palette = ['#8b5cf6','#3b82f6','#f59e0b','#22c55e','#ef4444','#ec4899','#14b8a6','#f97316'];
  
  charts.profs = new Chart(document.getElementById('chartProfs'), {{
    type: 'bar',
    data: {{
      labels: sorted.map(([p])=>abbreviateProfissional(p)),
      datasets: [{{
        label: 'Registros',
        data: sorted.map(([,c])=>c),
        backgroundColor: palette.map(c=>c+'99'),
        borderColor: palette, borderWidth:2, borderRadius:8
      }}]
    }},
    options: {{
      indexAxis:'y',
      responsive:true, maintainAspectRatio:true,
      plugins:{{ legend:{{ display:false }} }},
      scales:{{
        x:{{ grid:{{ color:'#334155' }}, ticks:{{ color:'#94a3b8' }} }},
        y:{{ grid:{{ display:false }}, ticks:{{ color:'#94a3b8', font:{{size:11}} }} }}
      }}
    }}
  }});
}}

// ────────────────────────────────────────────────────────────
// RENDER ALL
// ────────────────────────────────────────────────────────────
function renderAll() {{
  renderKPI();
  renderMap();
  renderCharts();
}}

// ────────────────────────────────────────────────────────────
// EVENT LISTENERS
// ────────────────────────────────────────────────────────────
document.getElementById('monthTabs').addEventListener('click', e => {{
  const btn = e.target.closest('.month-tab');
  if (!btn) return;
  document.querySelectorAll('.month-tab').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  currentSheet = btn.dataset.sheet;
  renderAll();
}});

document.getElementById('unitTabs').addEventListener('click', e => {{
  const btn = e.target.closest('.unit-tab');
  if (!btn) return;
  document.querySelectorAll('.unit-tab').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  currentUnit = btn.dataset.unit;
  renderAll();
}});

document.getElementById('searchInput').addEventListener('input', e => {{
  searchTerm = e.target.value;
  renderKPI();
  renderMap();
}});

function clearSearch() {{
  searchTerm = '';
  document.getElementById('searchInput').value = '';
  renderKPI();
  renderMap();
}}

// Initial render
renderAll();
</script>
</body>
</html>"""

with open('C:/Users/l5857/TCC - LETICIA/dashboard.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("dashboard.html created successfully!")
print(f"File size: {len(html)/1024:.1f} KB")
