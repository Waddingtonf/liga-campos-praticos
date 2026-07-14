import os
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

import sys
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8','utf8'):
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

# --- CONFIG ---
EXCEL_FILE = 'MAPEAMENTO DE CAMPO PRÁTICO (1).xlsx'
SHEET_ID = "1SFRlXmO_xmcD1HGnMN4LmALch2tkKGNlGFoOHtu5VYM"
SHEET_TABS = [
    "MARÇO",
    "FEVEREIRO (PLANEJAMENTO)",
    "JANEIRO(PLANEJAMENTO)",
]
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CRED_PATH = os.path.join(BASE_DIR, 'gen-lang-client-0965804770-1b0674d6e028.json')

# URL do site deployed (Vercel/Render) para limpar o cache automaticamente
LIVE_SITE_URL = "https://liga-campos-praticos.vercel.app" 

def sync():
    print(f"Lendo Excel: {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    print("Autenticando com Google Sheets...")
    scope = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    
    if os.environ.get('GCP_CREDENTIALS'):
        import json
        creds_dict = json.loads(os.environ.get('GCP_CREDENTIALS'))
        creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
    else:
        creds = Credentials.from_service_account_file(CRED_PATH, scopes=scope)
    
    client = gspread.authorize(creds)
    sh = client.open_by_key(SHEET_ID)
    
    for tab_name in SHEET_TABS:
        if tab_name not in wb.sheetnames:
            print(f"  ! Aba '{tab_name}' não encontrada no Excel. Pulando...")
            continue
            
        print(f"  -> Sincronizando aba: {tab_name}")
        excel_sheet = wb[tab_name]
        
        # 1. Tentar ler dados atuais do Google Sheets para preservar ALUNOS
        existing_students = {} # Key: (Unidade, Setor, Turno, Prof) -> Value: Alunos
        try:
            worksheet = sh.worksheet(tab_name)
            current_gs_data = worksheet.get_all_values()
            # Pular cabeçalhos (Linhas 0 e 1 no app.py)
            for i, row in enumerate(current_gs_data):
                if i < 2: continue
                # Colunas: 0:UNIT, 1:SETOR, 2:TURNO, 3:PROF, ..., 7:ALUNOS
                if len(row) > 7:
                    key = (row[0].strip().upper(), row[1].strip().upper(), row[2].strip().upper(), row[3].strip().upper())
                    students = row[7].strip()
                    if students and students not in ("None", ""):
                        existing_students[key] = students
        except gspread.exceptions.WorksheetNotFound:
            print(f"  + Aba '{tab_name}' não existe no Google Sheets. Será criada.")
            worksheet = sh.add_worksheet(title=tab_name, rows="100", cols="20")
        except Exception as e:
            print(f"  ! Aviso: Não foi possível ler dados atuais para preservar alunos: {e}")

        # 2. Extrair dados do Excel e fazer o MERGE
        final_data = []
        for r_idx, row in enumerate(excel_sheet.iter_rows(values_only=True)):
            row_clean = [str(cell) if cell is not None else "" for cell in row]
            
            # Garantir pelo menos 9 colunas para evitar IndexError
            if len(row_clean) < 9:
                row_clean += [""] * (9 - len(row_clean))
            
            first_cell = row_clean[0].strip().upper()
            
            if r_idx == 0:
                # Linha de título
                row_list = (row_clean + [""] * 10)[:10]
            elif first_cell == "UNIDADE":
                # Forçar cabeçalho unificado de 10 colunas
                row_list = [
                    "UNIDADE", "SETOR", "TURNO", "PROFISSIONAL", 
                    "CAPACIDADE", "OCUPAÇÃO", "CURSO", "ALUNOS", 
                    "TIPO DE OCUPAÇÃO", "PERÍODO"
                ]
            elif not row_clean[0] and not row_clean[1]:
                # Ignorar linhas completamente vazias
                row_list = [""] * 10
            else:
                # Mapeamento do Excel (9 colunas) para o Google Sheets (10 colunas)
                unidade, setor, turno, prof, cap_s, occ_s, curso, col7_val, col8_val = row_clean[:9]
                
                col7_val = col7_val.strip()
                col8_val = col8_val.strip()
                
                if col7_val.startswith('{') and col7_val.endswith('}'):
                    alunos_val = col7_val
                    tipo_val = ""
                else:
                    alunos_val = ""
                    tipo_val = col7_val
                
                periodo_val = col8_val
                
                # Preservação de alunos existentes no Sheets
                key = (unidade.strip().upper(), setor.strip().upper(), turno.strip().upper(), prof.strip().upper())
                if not (alunos_val.startswith('{') and alunos_val.endswith('}')):
                    if key in existing_students:
                        alunos_val = existing_students[key]
                
                row_list = [
                    unidade, setor, turno, prof, cap_s, occ_s, 
                    curso, alunos_val, tipo_val, periodo_val
                ]
            
            final_data.append(row_list)
            
        # 3. Limpar e atualizar dados
        worksheet.clear()
        if final_data:
            worksheet.update('A1', final_data)
            print(f"  ✓ {len(final_data)} linhas sincronizadas (Preservando coluna Aluno).")

    # --- REFRESH LIVE SITE CACHE ---
    if LIVE_SITE_URL:
        print(f"\nNotificando site em {LIVE_SITE_URL} para atualizar cache...")
        try:
            import requests
            refresh_url = f"{LIVE_SITE_URL.rstrip('/')}/api/refresh"
            resp = requests.post(refresh_url, timeout=10)
            if resp.status_code == 200:
                print("  ✓ Cache do site limpo com sucesso!")
            else:
                print(f"  ! Falha ao limpar cache: HTTP {resp.status_code}")
        except Exception as e:
            print(f"  ! Erro ao tentar limpar cache remoto: {e}")

if __name__ == "__main__":
    try:
        sync()
        print("\nSincronização concluída com sucesso!")
    except Exception as e:
        print(f"\nErro na sincronização: {e}")
