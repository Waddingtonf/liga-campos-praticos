import openpyxl
import json

wb = openpyxl.load_workbook('C:/Users/l5857/TCC - LETICIA/MAPEAMENTO DE CAMPO PRÁTICO (1).xlsx')
print('Sheets:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'\n=== Sheet: {sheet} ===')
    print(f'Dimensions: {ws.dimensions}')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 50:
            non_none = [x for x in row if x is not None]
            if non_none:
                print(f'Row {i+1}:', row)
        else:
            break
